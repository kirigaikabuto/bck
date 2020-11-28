from django.shortcuts import render, redirect, HttpResponse
from .forms import DataForm
from .models import Data, DataHistory
import openpyxl
import json
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
import numpy as np
from django.core.files.base import ContentFile


# Create your views here.
def home(request):
    if request.method == "POST":
        form = DataForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect("Home")
    form = DataForm()
    data = Data.objects.all()
    dataHistory = DataHistory.objects.all()
    ctx = {
        "form": form,
        "data": data,
        "dataHistory": dataHistory,
    }
    return render(request, "excapp/index.html", context=ctx)


def detail(request, id):
    data = Data.objects.get(pk=id)
    wb = openpyxl.load_workbook(data.file)
    worksheets = wb.worksheets[0]
    excel_data = list()
    for row in worksheets.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)

    context = {
        "excell": excel_data,
        "data": data,
    }
    return render(request, "excapp/detail.html", context=context)


def detail_info(request, id):
    data = Data.objects.get(pk=id)
    wb = openpyxl.load_workbook(data.file)
    worksheets = wb.worksheets[0]
    excel_data = list()
    for row in worksheets.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)
    return HttpResponse(json.dumps(excel_data, ensure_ascii=False))


@csrf_exempt
def upload_info(request):
    if request.method == "POST":
        excell = json.loads(request.read())
        excellData = excell["data"]
        id = int(excell["id"])
        data = Data.objects.get(pk=id)
        dataNumpy = np.array(excellData[1:])
        columns = excellData[0]
        df = pd.DataFrame(dataNumpy, columns=columns)
        from django.conf import settings
        import random
        fileName = data.name + str(random.randint(0, 1000))
        fileFullPath = settings.MEDIA_ROOT + "/documents/" + fileName + ".xlsx"
        df.to_excel(fileFullPath, index=False)
        file = open(fileFullPath, "rb+")
        newFile = ContentFile(file.read())
        newFile.name = fileName
        newData = Data.objects.create(name=fileName, file=newFile)
        # newData = Data.objects.create(name=fileName, file=newFile)
        newData.save()
        dataHistory = DataHistory.objects.create(parent=data, child=newData)
        dataHistory.save()
        print(dataHistory)
        return HttpResponse("ok")

    return HttpResponse("error")
